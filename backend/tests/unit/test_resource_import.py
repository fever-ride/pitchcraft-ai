"""Tests for resource Excel parsing logic (inlined to avoid motor dependency)."""
import io
import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook


def _make_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Inlined from backend.core.rag.resource_import to avoid motor import chain
KNOWN_COLUMNS = {
    "name", "type", "platform", "followers", "categories", "content_style",
    "audience_tags", "past_cpe", "tags", "pricing", "notes",
    "outlet_type", "beat", "service_type", "region",
    "placement_type", "location", "audience_reach",
}

HEADER_ALIASES = {
    "姓名": "name", "名称": "name", "资源名称": "name",
    "类型": "type", "平台": "platform",
    "粉丝数": "followers", "粉丝": "followers",
    "品类": "categories", "擅长品类": "categories",
    "内容风格": "content_style", "风格": "content_style",
    "受众": "audience_tags", "受众标签": "audience_tags",
    "历史cpe": "past_cpe", "cpe": "past_cpe",
    "标签": "tags", "报价": "pricing", "价格": "pricing", "备注": "notes",
    "媒体类型": "outlet_type", "跑线": "beat", "领域": "beat",
    "服务类型": "service_type", "区域": "region", "地区": "region",
    "广告类型": "placement_type", "位置": "location", "覆盖人群": "audience_reach",
}


class ImportParseResult:
    def __init__(self, resources, recognized, ignored):
        self.resources = resources
        self.recognized_columns = recognized
        self.ignored_columns = ignored


def parse_resource_excel(file_bytes: bytes) -> ImportParseResult:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return ImportParseResult([], [], [])

    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    mapped_headers = []
    recognized = []
    ignored = []
    for h in raw_headers:
        normalized = h.lower()
        if normalized in KNOWN_COLUMNS:
            mapped_headers.append(normalized)
            recognized.append(h)
        elif normalized in HEADER_ALIASES:
            mapped_headers.append(HEADER_ALIASES[normalized])
            recognized.append(h)
        else:
            mapped_headers.append("")
            if h:
                ignored.append(h)

    resources = []
    for row in rows[1:]:
        if not any(row):
            continue
        record = {}
        for idx, header in enumerate(mapped_headers):
            if not header:
                continue
            val = row[idx] if idx < len(row) else None
            if val is not None:
                record[header] = str(val).strip()
        if record.get("name"):
            record.setdefault("type", "kol")
            record.setdefault("platform", "")
            record.setdefault("tags", "")
            record.setdefault("categories", "")
            record.setdefault("content_style", "")
            record.setdefault("audience_tags", "")
            record.setdefault("past_cpe", "")
            for list_field in ("categories", "audience_tags", "tags"):
                val = record.get(list_field, "")
                if isinstance(val, str):
                    record[list_field] = [t.strip() for t in val.split(",") if t.strip()]
            resources.append(record)
    wb.close()
    return ImportParseResult(resources, recognized, ignored)


def _resource_to_text(r: dict) -> str:
    """Inlined copy of backend.core.rag.resource_import.resource_to_text.
    Keep in sync with the real function when that changes.
    """
    parts = [
        f"Name: {r.get('name', '')}",
        f"Type: {r.get('type', '')}",
    ]
    if r.get("tier"):
        parts.append(f"Tier: {r['tier']}")
    if r.get("platform"):
        parts.append(f"Platform: {r['platform']}")
    if r.get("followers"):
        parts.append(f"Followers: {r['followers']}")
    if r.get("categories"):
        cats = r["categories"]
        parts.append(f"Categories: {', '.join(cats) if isinstance(cats, list) else cats}")
    # Structured content style (v2) takes priority — but only if it has actual content
    cs = r.get("content_style_v2")
    style_parts = []
    if isinstance(cs, dict):
        if cs.get("production_level"):
            style_parts.append(f"production:{cs['production_level']}")
        if cs.get("persona_type"):
            style_parts.append(f"persona:{cs['persona_type']}")
        if cs.get("voice_style"):
            style_parts.append(f"voice:{cs['voice_style']}")
    if style_parts:
        parts.append(f"Content Style: {', '.join(style_parts)}")
    elif r.get("content_style"):
        parts.append(f"Content Style: {r['content_style']}")
    # Structured audience demographics takes priority over flat tags
    ad = r.get("audience_demographics")
    if isinstance(ad, dict):
        demo_parts = []
        if ad.get("age_range"):
            demo_parts.append(f"age:{ad['age_range']}")
        if ad.get("gender_skew"):
            demo_parts.append(ad["gender_skew"])
        if ad.get("city_tier"):
            demo_parts.append(ad["city_tier"])
        if ad.get("interest_tags"):
            demo_parts.extend(ad["interest_tags"])
        if demo_parts:
            parts.append(f"Audience: {', '.join(demo_parts)}")
    elif r.get("audience_tags"):
        tags = r["audience_tags"]
        parts.append(f"Audience: {', '.join(tags) if isinstance(tags, list) else tags}")
    if r.get("past_cpe"):
        parts.append(f"Past CPE: {r['past_cpe']}")
    if r.get("tags"):
        tags = r["tags"]
        parts.append(f"Tags: {', '.join(tags) if isinstance(tags, list) else tags}")
    if r.get("pricing"):
        parts.append(f"Pricing: {r['pricing']}")
    if r.get("notes"):
        parts.append(f"Notes: {r['notes']}")
    # Media-specific
    if r.get("outlet_type"):
        parts.append(f"Outlet: {r['outlet_type']}")
    if r.get("beat"):
        parts.append(f"Beat: {r['beat']}")
    # Vendor-specific
    if r.get("service_type"):
        parts.append(f"Service: {r['service_type']}")
    if r.get("region"):
        parts.append(f"Region: {r['region']}")
    # Placement-specific
    if r.get("placement_type"):
        parts.append(f"Placement: {r['placement_type']}")
    if r.get("location"):
        parts.append(f"Location: {r['location']}")
    if r.get("audience_reach"):
        parts.append(f"Reach: {r['audience_reach']}")
    return " | ".join(parts)


def test_parse_basic_rows():
    data = _make_xlsx([
        ["Name", "Type", "Platform", "Followers", "Tags", "Pricing", "Notes"],
        ["Alice", "kol", "小红书", "500000", "美妆,护肤", "5万/条", "Top tier"],
        ["Bob", "koc", "抖音", "50000", "生活方式", "5千/条", ""],
    ])
    result = parse_resource_excel(data)
    assert len(result.resources) == 2
    assert result.resources[0]["name"] == "Alice"
    assert result.resources[0]["type"] == "kol"
    assert result.resources[0]["platform"] == "小红书"
    assert result.resources[1]["name"] == "Bob"


def test_parse_skips_empty_rows():
    data = _make_xlsx([
        ["Name", "Type", "Platform"],
        ["Alice", "kol", "IG"],
        [None, None, None],
        ["Bob", "koc", "TikTok"],
    ])
    result = parse_resource_excel(data)
    assert len(result.resources) == 2


def test_parse_defaults_type_to_kol():
    data = _make_xlsx([
        ["Name", "Platform"],
        ["Charlie", "YouTube"],
    ])
    result = parse_resource_excel(data)
    assert result.resources[0]["type"] == "kol"


def test_parse_empty_file():
    data = _make_xlsx([])
    result = parse_resource_excel(data)
    assert result.resources == []


def test_column_recognition_english():
    data = _make_xlsx([
        ["Name", "Type", "Platform", "WeirdColumn", "Budget"],
        ["Alice", "kol", "IG", "foo", "100k"],
    ])
    result = parse_resource_excel(data)
    assert "Name" in result.recognized_columns
    assert "Type" in result.recognized_columns
    assert "Platform" in result.recognized_columns
    assert "WeirdColumn" in result.ignored_columns
    assert "Budget" in result.ignored_columns


def test_column_recognition_chinese_aliases():
    data = _make_xlsx([
        ["姓名", "平台", "粉丝数", "品类", "不认识的列"],
        ["Alice", "小红书", "50万", "美妆,护肤", "whatever"],
    ])
    result = parse_resource_excel(data)
    assert len(result.resources) == 1
    assert result.resources[0]["name"] == "Alice"
    assert result.resources[0]["platform"] == "小红书"
    assert result.resources[0]["followers"] == "50万"
    assert result.resources[0]["categories"] == ["美妆", "护肤"]
    assert "姓名" in result.recognized_columns
    assert "平台" in result.recognized_columns
    assert "粉丝数" in result.recognized_columns
    assert "品类" in result.recognized_columns
    assert "不认识的列" in result.ignored_columns


def test_column_recognition_mixed():
    data = _make_xlsx([
        ["Name", "报价", "UnknownA", "受众标签"],
        ["Bob", "5000", "x", "Z世代,女性"],
    ])
    result = parse_resource_excel(data)
    assert set(result.recognized_columns) == {"Name", "报价", "受众标签"}
    assert result.ignored_columns == ["UnknownA"]
    assert result.resources[0]["pricing"] == "5000"
    assert result.resources[0]["audience_tags"] == ["Z世代", "女性"]


# ---------------------------------------------------------------------------
# resource_to_text tests — verifying all fields make it into the search text
# ---------------------------------------------------------------------------

def test_resource_to_text_kol_full():
    """KOL with all common fields: everything should appear in the text."""
    r = {
        "name": "甜蜜生活Cindy", "type": "kol",
        "platform": "小红书", "tier": "top", "followers": "86万",
        "categories": ["美妆", "护肤"],
        "content_style": "精致生活分享",
        "audience_tags": ["20-30岁城市女性"],
        "past_cpe": "3.2%",
        "pricing": "8W-15W/篇",
        "notes": "档期提前8周",
    }
    text = _resource_to_text(r)
    assert "甜蜜生活Cindy" in text
    assert "kol" in text
    assert "小红书" in text
    assert "top" in text
    assert "86万" in text
    assert "美妆" in text
    assert "护肤" in text
    assert "精致生活分享" in text
    assert "20-30岁城市女性" in text
    assert "3.2%" in text
    assert "8W-15W/篇" in text
    assert "档期提前8周" in text


def test_resource_to_text_content_style_v2_takes_priority():
    """content_style_v2 (structured) should be used instead of content_style string."""
    r = {
        "name": "Test KOL", "type": "kol",
        "content_style": "这段文字不应该出现",
        "content_style_v2": {
            "production_level": "professional",
            "persona_type": "lifestyle",
            "voice_style": "casual",
        },
    }
    text = _resource_to_text(r)
    assert "professional" in text
    assert "lifestyle" in text
    assert "casual" in text
    assert "这段文字不应该出现" not in text


def test_resource_to_text_content_style_fallback():
    """Falls back to content_style string when v2 is absent."""
    r = {"name": "Test KOL", "type": "kol", "content_style": "接地气种草风"}
    text = _resource_to_text(r)
    assert "接地气种草风" in text


def test_resource_to_text_content_style_v2_empty_dict_falls_back():
    """content_style_v2 as empty dict should fall back to content_style string."""
    r = {
        "name": "Test KOL", "type": "kol",
        "content_style": "接地气种草风",
        "content_style_v2": {},
    }
    text = _resource_to_text(r)
    assert "接地气种草风" in text


def test_resource_to_text_audience_demographics_takes_priority():
    """Structured audience_demographics should be used instead of audience_tags list."""
    r = {
        "name": "Test KOL", "type": "kol",
        "audience_tags": ["这个标签不应该出现"],
        "audience_demographics": {
            "age_range": "25-35",
            "gender_skew": "female",
            "city_tier": "tier1",
            "interest_tags": ["美妆", "健身"],
        },
    }
    text = _resource_to_text(r)
    assert "25-35" in text
    assert "female" in text
    assert "tier1" in text
    assert "美妆" in text
    assert "健身" in text
    assert "这个标签不应该出现" not in text


def test_resource_to_text_audience_tags_fallback():
    """Falls back to audience_tags list when demographics absent."""
    r = {
        "name": "Test KOL", "type": "kol",
        "audience_tags": ["Z世代", "学生党"],
    }
    text = _resource_to_text(r)
    assert "Z世代" in text
    assert "学生党" in text


def test_resource_to_text_media():
    """Media resource: outlet_type and beat must appear."""
    r = {
        "name": "36氪", "type": "media",
        "outlet_type": "online",
        "beat": "科技/创投",
        "pricing": "发稿15000-35000",
    }
    text = _resource_to_text(r)
    assert "36氪" in text
    assert "media" in text
    assert "online" in text
    assert "科技/创投" in text
    assert "15000" in text
    # KOL-only fields should not appear
    assert "Platform:" not in text
    assert "Followers:" not in text


def test_resource_to_text_vendor():
    """Vendor resource: service_type and region must appear."""
    r = {
        "name": "禾木创意", "type": "vendor",
        "service_type": "活动策划/整合执行",
        "region": "上海/华东",
        "notes": "擅长大型品牌活动",
    }
    text = _resource_to_text(r)
    assert "禾木创意" in text
    assert "vendor" in text
    assert "活动策划" in text
    assert "上海/华东" in text
    assert "擅长大型品牌活动" in text


def test_resource_to_text_placement():
    """Placement resource: placement_type, location, audience_reach must appear."""
    r = {
        "name": "来福士广场LED", "type": "placement",
        "placement_type": "OOH",
        "location": "上海静安区",
        "audience_reach": "日均15万人次",
    }
    text = _resource_to_text(r)
    assert "OOH" in text
    assert "上海静安区" in text
    assert "日均15万人次" in text


def test_resource_to_text_missing_optional_fields_no_crash():
    """Bare minimum doc (only name + type) should not crash."""
    text = _resource_to_text({"name": "Minimal", "type": "kol"})
    assert "Minimal" in text
    assert "kol" in text


def test_resource_to_text_categories_as_list():
    r = {"name": "X", "type": "kol", "categories": ["美妆", "护肤", "香氛"]}
    text = _resource_to_text(r)
    assert "美妆" in text
    assert "护肤" in text
    assert "香氛" in text


def test_resource_to_text_categories_as_string():
    """categories stored as plain string (legacy) should still appear."""
    r = {"name": "X", "type": "kol", "categories": "美妆,护肤"}
    text = _resource_to_text(r)
    assert "美妆,护肤" in text


def test_resource_to_text_empty_categories_excluded():
    """Empty categories list should not add a blank 'Categories:' segment."""
    r = {"name": "X", "type": "kol", "categories": []}
    text = _resource_to_text(r)
    assert "Categories:" not in text


def test_resource_to_text_platform_excluded_when_absent():
    """Platform absent → no 'Platform:' label in text (media/vendor have no platform)."""
    r = {"name": "虎嗅", "type": "media"}
    text = _resource_to_text(r)
    assert "Platform:" not in text


def test_resource_to_text_separator_is_pipe():
    """Fields are joined with ' | ' separator — important for embedding tokenization."""
    r = {"name": "Alice", "type": "kol", "platform": "抖音"}
    text = _resource_to_text(r)
    assert " | " in text


def test_alias_case_insensitive():
    """Header alias matching should be case-insensitive."""
    data = _make_xlsx([
        ["姓名", "PLATFORM", "Followers"],
        ["Eve", "IG", "10k"],
    ])
    result = parse_resource_excel(data)
    assert len(result.resources) == 1
    assert result.resources[0]["name"] == "Eve"
    assert result.resources[0]["platform"] == "IG"
    assert result.resources[0]["followers"] == "10k"


def test_all_ignored_columns():
    """File with no recognized columns still returns empty resources."""
    data = _make_xlsx([
        ["Foo", "Bar", "Baz"],
        ["a", "b", "c"],
    ])
    result = parse_resource_excel(data)
    assert result.resources == []
    assert set(result.ignored_columns) == {"Foo", "Bar", "Baz"}
    assert result.recognized_columns == []


def test_list_normalization_in_parse():
    """categories and audience_tags are split from comma-separated string to list."""
    data = _make_xlsx([
        ["Name", "Categories", "Audience_tags", "Tags"],
        ["Frank", "美妆,护肤,彩妆", "Z世代,女性", "头部,种草"],
    ])
    result = parse_resource_excel(data)
    r = result.resources[0]
    assert r["categories"] == ["美妆", "护肤", "彩妆"]
    assert r["audience_tags"] == ["Z世代", "女性"]
    assert r["tags"] == ["头部", "种草"]


def test_empty_list_fields_default():
    """Empty categories/audience_tags default to empty list."""
    data = _make_xlsx([
        ["Name", "Type"],
        ["Grace", "media"],
    ])
    result = parse_resource_excel(data)
    r = result.resources[0]
    assert r["categories"] == []
    assert r["audience_tags"] == []
    assert r["tags"] == []
