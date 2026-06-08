"""Resource Excel import: parse .xlsx, create DB records, embed, upsert to Pinecone."""
import asyncio
import io
import json
import logging
import re as _re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

from openpyxl import load_workbook

from backend.core.database.connection import get_database
from backend.core.database.repositories.resources import ResourceRepository
from backend.core.models.resource import ResourceStatus, normalize_platform, parse_follower_count, resource_namespace
from backend.core.rag.embedder import embed_texts
from backend.core.rag.indexer import upsert_vectors

VALID_TYPES = {"kol", "koc", "media", "vendor", "placement"}

# Columns that are definitively structural noise — skip before LLM inference, never map to schema.
KNOWN_IGNORE_COLUMNS = {
    "序号", "no", "no.", "#", "编号", "id", "index", "行号",
    "备选", "是否合作", "合作状态", "状态", "是否可用",
}

KNOWN_COLUMNS = {
    "name", "type", "tier", "platform", "followers", "categories",
    "content_style", "production_level", "persona_type", "voice_style",
    "audience_tags", "age_range", "gender_skew", "city_tier", "interest_tags",
    "past_cpe", "tags", "pricing", "notes",
    "outlet_type", "beat", "service_type", "region",
    "placement_type", "location", "audience_reach",
    "profile_url", "contact",
}

HEADER_ALIASES = {
    # name
    "姓名": "name",
    "名称": "name",
    "资源名称": "name",
    "达人昵称": "name",
    "达人名称": "name",
    "昵称": "name",
    "博主昵称": "name",
    "博主名称": "name",
    "账号名称": "name",
    "账号": "name",
    "媒体名称": "name",
    "供应商名称": "name",
    "kol名称": "name",
    # type — use unambiguous aliases only; "类型" alone almost always means content category in real KOL sheets
    "资源类型": "type",
    "达人类型": "type",
    "kol类型": "type",
    # categories — "类型" in practice means content genre (摄影/生活方式), not kol/koc/media
    "类型": "categories",
    "内容类型": "categories",
    # tier
    "层级": "tier",
    "达人层级": "tier",
    "体量": "tier",
    "体量/等级": "tier",
    "等级": "tier",
    "kol等级": "tier",
    # platform
    "平台": "platform",
    "所在平台": "platform",
    "主要平台": "platform",
    "所属平台": "platform",
    # followers
    "粉丝数": "followers",
    "粉丝": "followers",
    "粉丝量": "followers",
    "关注数": "followers",
    "粉丝数量": "followers",
    # categories
    "品类": "categories",
    "擅长品类": "categories",
    "擅长类目": "categories",
    "垂类": "categories",
    "内容品类": "categories",
    "行业": "categories",
    # content style
    "内容风格": "content_style",
    "风格": "content_style",
    "内容调性": "content_style",
    # content style v2 dimensions
    "制作水平": "production_level",
    "人设类型": "persona_type",
    "表达风格": "voice_style",
    # audience
    "受众": "audience_tags",
    "受众标签": "audience_tags",
    "目标受众": "audience_tags",
    "粉丝画像": "audience_tags",
    "年龄段": "age_range",
    "受众年龄": "age_range",
    "粉丝年龄": "age_range",
    "性别倾向": "gender_skew",
    "粉丝性别": "gender_skew",
    "城市级别": "city_tier",
    "城市分布": "city_tier",
    "兴趣标签": "interest_tags",
    "粉丝兴趣": "interest_tags",
    # performance
    "历史cpe": "past_cpe",
    "cpe": "past_cpe",
    "互动率": "past_cpe",
    # pricing
    "标签": "tags",
    "报价": "pricing",
    "价格": "pricing",
    "参考报价": "pricing",
    "合作报价": "pricing",
    # notes
    "备注": "notes",
    "说明": "notes",
    # profile url (homepage / social profile link)
    "链接": "profile_url",
    "主页链接": "profile_url",
    "主页": "profile_url",
    "账号链接": "profile_url",
    "账号主页": "profile_url",
    "profile": "profile_url",
    "url": "profile_url",
    # contact (email / phone — non-url contact info)
    "联系方式": "contact",
    "联系人": "contact",
    "邮箱": "contact",
    "电话": "contact",
    "微信": "contact",
    # media-specific
    "媒体类型": "outlet_type",
    "跑线": "beat",
    "领域": "beat",
    # vendor-specific
    "服务类型": "service_type",
    "区域": "region",
    "地区": "region",
    "服务区域": "region",
    # placement-specific
    "广告类型": "placement_type",
    "位置": "location",
    "覆盖人群": "audience_reach",
}


SCHEMA_FIELD_DESCRIPTIONS: dict[str, str] = {
    "name":             "资源名称（KOL/媒体/供应商的名字或账号名）",
    "type":             "资源类型：kol, koc, media, vendor, placement",
    "tier":             "达人层级：top头部, mid中腰, tail尾部, koc",
    "platform":         "所在平台：小红书/抖音/微博/bilibili等",
    "followers":        "粉丝数量",
    "categories":       "内容品类/垂类/行业",
    "content_style":    "内容风格（自由文本描述）",
    "production_level": "内容制作水平：high精良/medium标准/low低成本",
    "persona_type":     "人设类型：expert专家/relatable平民/aspirational向往/entertaining娱乐",
    "voice_style":      "表达风格：educational科普/conversational对话/emotional情感/humorous幽默",
    "audience_tags":    "受众标签（如'年轻女性,美妆爱好者'）",
    "age_range":        "受众年龄段（如'18-25'）",
    "gender_skew":      "粉丝性别倾向：female/male/balanced",
    "city_tier":        "城市级别：tier_1一线/tier_2_3二三线/all全国",
    "interest_tags":    "受众兴趣标签",
    "past_cpe":         "历史互动成本/CPE数据",
    "tags":             "资源标签（自定义标签）",
    "pricing":          "报价/合作价格",
    "notes":            "备注说明",
    "outlet_type":      "媒体类型（media专用）：newspaper/magazine/online/TV",
    "beat":             "跑线/报道领域（media专用）",
    "service_type":     "服务类型（vendor专用）：event/photography/production",
    "region":           "服务区域/地区",
    "placement_type":   "广告形式（placement专用）：OOH/elevator/cinema",
    "location":         "广告位地点（placement专用）",
    "audience_reach":   "覆盖人群规模（placement专用）",
    "profile_url":      "主页链接/账号链接（http/https开头的URL）",
    "contact":          "联系方式（邮箱、电话、微信等非URL联系信息）",
}


class ColumnInference(BaseModel):
    header: str = Field(description="原始列名")
    field: str | None = Field(default=None, description="匹配到的schema字段名，无匹配时为null")
    confidence: str = Field(default="medium", description="置信度：high/medium/low")
    reason: str = Field(default="", description="简短说明")


class ColumnInferenceResult(BaseModel):
    mappings: list[ColumnInference]


TIER_NORMALIZER: dict[str, str] = {
    # English canonical
    "top": "top", "mid": "mid", "tail": "tail", "koc": "koc",
    # Chinese
    "头部": "top", "顶部": "top", "头部达人": "top",
    "腰部": "mid", "中腰": "mid", "中部": "mid", "腰部达人": "mid",
    "尾部": "tail", "长尾": "tail", "尾部达人": "tail",
    # Level-based (lv1-2=top, lv3-4=mid, lv5-6=tail)
    "lv1": "top", "lv2": "top",
    "lv3": "mid", "lv4": "mid",
    "lv5": "tail", "lv6": "tail",
    # Grade-based
    "s级": "top", "s": "top",
    "a级": "mid", "a": "mid",
    "b级": "tail", "b": "tail",
    "c级": "tail", "c": "tail",
    # Size-based
    "大v": "top", "头部kol": "top",
    "中kol": "mid", "中腰kol": "mid",
    "小v": "tail", "尾部kol": "tail",
}


def _normalize_tier(raw: str) -> str | None:
    """Normalize tier value from various conventions to canonical top/mid/tail/koc."""
    if not raw:
        return None
    return TIER_NORMALIZER.get(raw.strip().lower())


async def infer_unrecognized_headers(headers: list[str]) -> list[dict]:
    """Use LLM to infer schema field mappings for headers not in our alias table."""
    if not headers:
        return []

    from langchain_core.messages import HumanMessage, SystemMessage
    from backend.core.agents.llm import invoke_llm_structured

    fields_text = "\n".join(f"  {k}: {v}" for k, v in SCHEMA_FIELD_DESCRIPTIONS.items())
    messages = [
        SystemMessage(content="你是数据字段映射助手。将Excel表头映射到资源库schema字段。只返回有意义的映射，无法匹配的返回null。"),
        HumanMessage(content=(
            f"将以下未识别的Excel表头映射到最合适的schema字段。\n\n"
            f"可用字段：\n{fields_text}\n\n"
            f"待映射表头（JSON数组）：{json.dumps(headers, ensure_ascii=False)}"
        )),
    ]
    result = await invoke_llm_structured(messages, output_schema=ColumnInferenceResult, temperature=0)
    return [m.model_dump() for m in result.mappings]


async def preview_import(file_bytes: bytes) -> dict:
    """Parse Excel headers and return full column mapping analysis.

    Applies static alias lookup first; calls LLM for any remaining unrecognized headers.
    Does NOT write to DB or Pinecone — safe to call without side effects.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {"row_count": 0, "recognized": [], "inferred": [], "ignored": []}

    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    row_count = len(rows) - 1

    recognized: list[dict] = []
    ignored: list[dict] = []
    unrecognized: list[str] = []

    for h in raw_headers:
        if not h:
            continue
        norm = h.lower()
        if norm in KNOWN_IGNORE_COLUMNS:
            ignored.append({"raw": h, "reason": "结构性序号列，自动忽略"})
        elif norm in KNOWN_COLUMNS:
            recognized.append({"raw": h, "mapped_to": norm, "source": "exact"})
        elif norm in HEADER_ALIASES:
            recognized.append({"raw": h, "mapped_to": HEADER_ALIASES[norm], "source": "alias"})
        else:
            unrecognized.append(h)

    inferred: list[dict] = []

    if unrecognized:
        llm_results = await infer_unrecognized_headers(unrecognized)
        for item in llm_results:
            if item.get("field"):
                inferred.append({
                    "raw": item["header"],
                    "mapped_to": item["field"],
                    "confidence": item.get("confidence", "medium"),
                    "reason": item.get("reason", ""),
                })
            else:
                ignored.append({
                    "raw": item["header"],
                    "reason": item.get("reason", "无对应字段"),
                })

    return {
        "row_count": row_count,
        "recognized": recognized,   # 静态规则直接识别
        "inferred": inferred,        # LLM 推断，需用户确认
        "ignored": ignored,          # 无法映射，将被丢弃
    }


class ImportParseResult:
    def __init__(self, resources: list[dict], recognized: list[str], ignored: list[str]):
        self.resources = resources
        self.recognized_columns = recognized
        self.ignored_columns = ignored


_PLATFORM_SPLIT_RE = _re.compile(r'[+、/，,]+')
_PLATFORM_FOLLOWER_RE = _re.compile(
    r'(小红书|抖音|微博|微信|b站|bilibili|快手|youtube|instagram)([\d.]+(?:万|k|m)?)',
    _re.IGNORECASE,
)


def _build_platforms(raw_platform: str, raw_followers: str, raw_url: str) -> list[dict]:
    """Convert raw string fields into a list of PlatformEntry dicts."""
    platform_names = [p.strip() for p in _PLATFORM_SPLIT_RE.split(raw_platform) if p.strip()]
    if not platform_names:
        return []

    # Try to parse per-platform followers e.g. "抖音88万+小红书32万"
    per_platform: dict[str, int] = {}
    if raw_followers:
        for m in _PLATFORM_FOLLOWER_RE.finditer(raw_followers):
            fc = parse_follower_count(m.group(2))
            if fc:
                per_platform[m.group(1).lower()] = fc

    total_count = parse_follower_count(raw_followers)

    entries = []
    for i, pname in enumerate(platform_names):
        fc = per_platform.get(pname.lower())
        if fc is None and len(platform_names) == 1:
            fc = total_count
        entries.append({
            "name": pname,
            "followers_raw": raw_followers if len(platform_names) == 1 else None,
            "followers_count": fc,
            "profile_url": raw_url if i == 0 else None,
        })
    return entries


def parse_resource_excel(
    file_bytes: bytes,
    override_mapping: dict[str, str] | None = None,
) -> ImportParseResult:
    """Parse xlsx bytes into list of resource dicts with column recognition feedback.

    override_mapping: {raw_header: field_name} confirmed by user after preview.
                      Use "ignore" as field_name to explicitly skip a column.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return ImportParseResult([], [], [])

    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    raw_headers_lower = {h.lower() for h in raw_headers if h}

    # Map headers: exact match → alias lookup → user override → ignored
    mapped_headers = []
    recognized = []
    ignored = []
    for h in raw_headers:
        normalized = h.lower()
        if normalized in KNOWN_IGNORE_COLUMNS:
            mapped_headers.append("")   # silently drop — structural noise
            ignored.append(h)
        elif normalized in KNOWN_COLUMNS:
            mapped_headers.append(normalized)
            recognized.append(h)
        elif normalized in HEADER_ALIASES:
            mapped_headers.append(HEADER_ALIASES[normalized])
            recognized.append(h)
        elif override_mapping and h in override_mapping:
            field = override_mapping[h]
            if field and field.lower() != "ignore":
                mapped_headers.append(field)
                recognized.append(h)
            else:
                mapped_headers.append("")   # user explicitly ignored
        else:
            mapped_headers.append("")
            if h:
                ignored.append(h)

    def _is_repeat_header(row: tuple) -> bool:
        """Return True if this row looks like a repeated header row (e.g. mid-sheet section dividers)."""
        non_empty = [str(v).strip().lower() for v in row if v is not None and str(v).strip()]
        if not non_empty:
            return False
        matches = sum(1 for v in non_empty if v in raw_headers_lower)
        return matches >= min(3, max(2, len(non_empty) // 2))

    resources = []
    for row in rows[1:]:
        if not any(row):
            continue
        if _is_repeat_header(row):
            continue
        record = {}
        for idx, header in enumerate(mapped_headers):
            if not header:
                continue
            val = row[idx] if idx < len(row) else None
            if val is not None:
                record[header] = str(val).strip()
        if record.get("name"):
            record.setdefault("type", "kol")
            record.setdefault("tags", "")
            record.setdefault("categories", "")
            record.setdefault("content_style", "")
            record.setdefault("audience_tags", "")
            record.setdefault("past_cpe", "")
            # Build platforms list from raw fields, then remove flat fields
            raw_platform = record.pop("platform", "")
            raw_followers = record.pop("followers", "")
            raw_url = record.pop("profile_url", "")
            record.pop("followers_count", None)       # remove old computed field

            platforms = _build_platforms(raw_platform, raw_followers, raw_url)
            record["platforms"] = platforms
            record["primary_platform"] = normalize_platform(platforms[0]["name"]) if platforms else ""
            per_platform_total = sum(p["followers_count"] for p in platforms if p.get("followers_count")) or None
            record["total_followers_count"] = per_platform_total or parse_follower_count(raw_followers)
            record["status"] = ResourceStatus.ACTIVE.value
            record["last_verified_at"] = datetime.utcnow()
            for list_field in ("categories", "audience_tags", "tags", "interest_tags"):
                val = record.get(list_field, "")
                if isinstance(val, str):
                    record[list_field] = [t.strip() for t in val.split(",") if t.strip()]
            # Assemble structured content_style_v2 from individual columns
            cs_v2 = {}
            for cs_field in ("production_level", "persona_type", "voice_style"):
                val = record.pop(cs_field, None)
                if val:
                    cs_v2[cs_field] = val
            if cs_v2:
                record["content_style_v2"] = cs_v2
            # Assemble structured audience_demographics from individual columns
            ad = {}
            for ad_field in ("age_range", "gender_skew", "city_tier"):
                val = record.pop(ad_field, None)
                if val:
                    ad[ad_field] = val
            interest = record.pop("interest_tags", [])
            if interest:
                ad["interest_tags"] = interest
            if ad:
                record["audience_demographics"] = ad
            # Normalize tier — accept English, Chinese, and level-based values
            tier_val = record.get("tier", "")
            record["tier"] = _normalize_tier(tier_val) if tier_val else None
            if not record["tier"]:
                record.pop("tier", None)
            resources.append(record)

    wb.close()
    return ImportParseResult(resources, recognized, ignored)


def resource_to_text(r: dict) -> str:
    """Convert resource dict to searchable text for embedding."""
    parts = [
        f"Name: {r.get('name', '')}",
        f"Type: {r.get('type', '')}",
    ]
    if r.get("tier"):
        parts.append(f"Tier: {r['tier']}")
    if r.get("platforms"):
        for p in r["platforms"]:
            line = f"Platform: {p['name']}"
            if p.get("followers_count"):
                line += f" ({p['followers_count']:,} followers)"
            parts.append(line)
    elif r.get("primary_platform"):
        parts.append(f"Platform: {r['primary_platform']}")
    if r.get("categories"):
        cats = r["categories"]
        parts.append(f"Categories: {', '.join(cats) if isinstance(cats, list) else cats}")
    # Structured content style (v2) takes priority — but only if it has actual content
    cs = r.get("content_style_v2")
    style_parts = []
    if isinstance(cs, dict):
        if cs.get("production_level"):
            style_parts.append(f"production:{cs['production_level']}")
        if cs.get("persona_type"):
            style_parts.append(f"persona:{cs['persona_type']}")
        if cs.get("voice_style"):
            style_parts.append(f"voice:{cs['voice_style']}")
    if style_parts:
        parts.append(f"Content Style: {', '.join(style_parts)}")
    elif r.get("content_style"):
        parts.append(f"Content Style: {r['content_style']}")
    # Structured audience demographics
    ad = r.get("audience_demographics")
    if isinstance(ad, dict):
        demo_parts = []
        if ad.get("age_range"):
            demo_parts.append(f"age:{ad['age_range']}")
        if ad.get("gender_skew"):
            demo_parts.append(ad["gender_skew"])
        if ad.get("city_tier"):
            demo_parts.append(ad["city_tier"])
        if ad.get("interest_tags"):
            demo_parts.extend(ad["interest_tags"])
        if demo_parts:
            parts.append(f"Audience: {', '.join(demo_parts)}")
    elif r.get("audience_tags"):
        tags = r["audience_tags"]
        parts.append(f"Audience: {', '.join(tags) if isinstance(tags, list) else tags}")
    if r.get("past_cpe"):
        parts.append(f"Past CPE: {r['past_cpe']}")
    if r.get("tags"):
        tags = r["tags"]
        parts.append(f"Tags: {', '.join(tags) if isinstance(tags, list) else tags}")
    if r.get("pricing"):
        parts.append(f"Pricing: {r['pricing']}")
    if r.get("notes"):
        parts.append(f"Notes: {r['notes']}")
    # Media-specific
    if r.get("outlet_type"):
        parts.append(f"Outlet: {r['outlet_type']}")
    if r.get("beat"):
        parts.append(f"Beat: {r['beat']}")
    # Vendor-specific
    if r.get("service_type"):
        parts.append(f"Service: {r['service_type']}")
    if r.get("region"):
        parts.append(f"Region: {r['region']}")
    # Placement-specific
    if r.get("placement_type"):
        parts.append(f"Placement: {r['placement_type']}")
    if r.get("location"):
        parts.append(f"Location: {r['location']}")
    if r.get("audience_reach"):
        parts.append(f"Reach: {r['audience_reach']}")
    return " | ".join(parts)


async def refresh_resource_embedding(doc: dict, client_id: str):
    """Re-embed a single resource and upsert to Pinecone. Used after any field update."""
    rtype = doc.get("type", "kol").lower()
    if rtype not in VALID_TYPES:
        rtype = "kol"
    ns = resource_namespace(rtype, client_id)
    text = resource_to_text(doc)
    embeddings = await embed_texts([text])
    resource_id = str(doc.get("_id", ""))
    extra_meta = {
        "name": doc.get("name", ""),
        "type": doc.get("type", rtype),
        "tier": doc.get("tier", ""),
        "platforms": [normalize_platform(p["name"]) for p in doc.get("platforms", []) if p.get("name")],
        "primary_platform": doc.get("primary_platform", ""),
        "status": doc.get("status", ResourceStatus.ACTIVE.value),
        "total_followers_count": doc.get("total_followers_count") or 0,
        "tags": ", ".join(doc.get("tags", [])) if isinstance(doc.get("tags"), list) else doc.get("tags", ""),
    }
    upsert_vectors(ns, resource_id, [text], embeddings, extra_metadata=[extra_meta])


async def import_resources(
    file_bytes: bytes,
    client_id: str,
    override_mapping: dict[str, str] | None = None,
) -> dict:
    """Full import pipeline: parse → dedup → DB (bulk) → embed → Pinecone (grouped by type)."""
    parse_result = parse_resource_excel(file_bytes, override_mapping=override_mapping)
    resources = parse_result.resources
    if not resources:
        return {
            "imported": 0,
            "error": "No valid rows found",
            "recognized_columns": parse_result.recognized_columns,
            "ignored_columns": parse_result.ignored_columns,
        }

    db = await get_database()
    repo = ResourceRepository(db)

    # Dedup: skip resources whose name already exists for this client
    existing_names = await repo.get_names_set(client_id)
    new_resources = [r for r in resources if r.get("name", "").lower() not in existing_names]
    skipped = len(resources) - len(new_resources)

    if not new_resources:
        return {
            "imported": 0,
            "skipped": skipped,
            "reason": "All resources already exist",
            "recognized_columns": parse_result.recognized_columns,
            "ignored_columns": parse_result.ignored_columns,
        }

    # Bulk insert — motor sets _id on each dict in-place
    for r in new_resources:
        r["client_id"] = client_id
    await repo.create_many(new_resources)

    # Group by type for namespace-specific embedding + upsert
    by_type: dict[str, list[dict]] = defaultdict(list)
    for r in new_resources:
        rtype = r.get("type", "kol").lower()
        if rtype not in VALID_TYPES:
            rtype = "kol"
        by_type[rtype].append(r)

    namespaces_used = []
    for rtype, group in by_type.items():
        ns = resource_namespace(rtype, client_id)
        texts = [resource_to_text(r) for r in group]
        embeddings = await embed_texts(texts)

        # Use MongoDB _id as vector ID — eliminates re-import collision
        vector_ids = [str(r["_id"]) for r in group]

        extra_metadata = []
        for r in group:
            tags = r.get("tags", [])
            meta = {
                "name": r.get("name", ""),
                "type": r.get("type", rtype),
                "tier": r.get("tier", ""),
                "platforms": [normalize_platform(p["name"]) for p in r.get("platforms", []) if p.get("name")],
                "primary_platform": r.get("primary_platform", ""),
                "status": r.get("status", ResourceStatus.ACTIVE.value),
                "total_followers_count": r.get("total_followers_count") or 0,
                "tags": ", ".join(tags) if isinstance(tags, list) else tags,
            }
            extra_metadata.append(meta)

        upsert_vectors(
            namespace=ns,
            file_id=f"import_{client_id}_{rtype}",
            chunks=texts,
            embeddings=embeddings,
            extra_metadata=extra_metadata,
            ids=vector_ids,
        )
        namespaces_used.append(ns)

    return {
        "imported": len(new_resources),
        "skipped": skipped,
        "by_type": {k: len(v) for k, v in by_type.items()},
        "namespaces": namespaces_used,
        "recognized_columns": parse_result.recognized_columns,
        "ignored_columns": parse_result.ignored_columns,
    }


async def repair_resource_embeddings(client_id: str) -> dict:
    """Re-embed ALL resources for a client and upsert to Pinecone.

    Fixes orphaned records: resources present in MongoDB but missing from Pinecone
    due to a failed import (embed/Pinecone error after DB insert).
    Safe to run repeatedly — upsert is idempotent.
    """
    db = await get_database()
    repo = ResourceRepository(db)

    # Include all statuses so inactive resources also get their vectors repaired
    docs = await repo.find({"client_id": client_id}, limit=5000)
    if not docs:
        logger.info(f"repair_resource_embeddings: no resources found for {client_id}")
        return {"repaired": 0, "errors": 0, "total": 0}

    repaired = 0
    errors = 0
    for doc in docs:
        try:
            await refresh_resource_embedding(doc, client_id)
            repaired += 1
        except Exception as e:
            logger.error(
                f"repair_resource_embeddings: failed for {doc.get('name', doc.get('_id'))}: {e}"
            )
            errors += 1

    logger.info(
        f"repair_resource_embeddings: {client_id} → repaired={repaired}, errors={errors}"
    )
    return {"repaired": repaired, "errors": errors, "total": len(docs)}


def make_import_task():
    """Deferred import to avoid circular import at module load time."""
    from backend.core.tasks import celery_app

    @celery_app.task(
        bind=True,
        name="resource_import.import_resources_task",
        max_retries=1,
        default_retry_delay=30,
        time_limit=600,
    )
    def import_resources_task(
        self,
        storage_path: str,
        client_id: str,
        override_mapping: dict | None = None,
    ) -> dict:
        """Celery task: bulk Excel import — runs embed + Pinecone upsert in background."""
        try:
            file_bytes = Path(storage_path).read_bytes()
            return asyncio.run(import_resources(file_bytes, client_id, override_mapping=override_mapping))
        except Exception as exc:
            logger.error(f"import_resources_task failed for {client_id}: {exc}")
            raise self.retry(exc=exc)
        finally:
            Path(storage_path).unlink(missing_ok=True)

    return import_resources_task


try:
    import_resources_task = make_import_task()
except ImportError:
    import_resources_task = None  # Celery not available (e.g. test environment)
